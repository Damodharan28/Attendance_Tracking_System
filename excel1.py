from openpyxl import Workbook
from openpyxl import load_workbook
from django.utils import timezone
from tablib import Dataset
# from ATS_app.models import ATTENDANCE_INFO

# # Create a new Excel workbook
# # wb = Workbook()

# # Select the active worksheet
# # ws = wb.active

# current_date = timezone.now().date()
# day_week = current_date.weekday()
# print(day_week)
# days={
#      0 : "MONDAY",
#      1 : "TUESDAY",
#      2 : "WEDNESDAY",
#      3 : "THURSDAY",
#      4 : "FRIDAY"
# }

# dataset = Dataset()
# imported_data = dataset.load(open('excels/timetable.xlsx', 'rb').read(),format='xlsx')

# subjects =[]
# for data in imported_data:
#     if data[0] == days[day_week]:
#         for i in range(1,9):
#             subjects.append(data[i])
# print(subjects)
# wb = load_workbook('attendance.xlsx')

# # Iterate over all sheets in the workbook
# for sheet in wb.sheetnames:
#     ws = wb[sheet]
#     # Delete all rows in the sheet
#     ws.delete_rows(1, ws.max_row)

# # Define the headers
# wb.append(["ATTENDANCE_ID","DATE","HOUR","SUBJECT_ID","STATUS"])

# all_objects = ATTENDANCE_INFO.objects.all()

# # Iterate over the objects and access their fields
# for obj in all_objects:
#     for i in range(1,9):
#         wb.append(obj.ATTENDANCE_ID,current_date,i,subjects[i-1],"PRESENT") 

# wb.save('attendance.xlsx')
# print("created")
# with open('excels/attendance.xlsx', 'rb') as excel_file:
#         response = HttpResponse(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename=attendance.xlsx'
#         return response


# Save the workbook to a file
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import random

# # Load the existing Excel file
# excel_file = "attendance.xlsx"
# wb = load_workbook(excel_file)
# ws = wb.active

# ws.append(["ATTENDANCE_ID", "DATE" ,"HOUR", "SUBJECT_ID", "STATUS"])

# # Sample data generation (replace this with your actual data generation code)
# # student_id=20001
# attendance_id=50000
# subjects = ["JCS1001", "JCS1002", "JCS1003", "JCS1004", "JCS1005", "JCS1006", "JCS1007", "JCS1008", "JCS1009", "JCS1010"]
# for i in range(50):
#     # student_id = student_id + i
#     attendance_id = attendance_id + i
#     for hour in range(1, 9):
#         for subject_id_num in range(1, 11):
#             # attendance_id = 50000 + (i * 10 + hour * 10 + subject_id_num)
#             date = datetime.now().strftime("%Y-%m-%d")
#             hour_str = str(hour)

#             # Select 8 random subjects
#             random_subjects = random.sample(subjects, 8)

#             for hour, subject_id in enumerate(random_subjects, start=1):
#                 hour_str = str(hour)
#                 status = random.choice(["PRESENT", "ABSENT","ON DUTY"])

#                 # Append data to the worksheet
#                 row_data = [attendance_id, date, hour_str, subject_id, status]
#                 ws.append(row_data)
            

# # Save the workbook back to the Excel file
# wb.save(excel_file)
date = datetime.now().strftime("%Y-%m-%d")
print(date)