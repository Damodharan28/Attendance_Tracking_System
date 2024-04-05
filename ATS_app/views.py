from django.shortcuts import render,redirect
from django.http import HttpResponse
from django.http import JsonResponse
import random
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.db.utils import IntegrityError
from django.db.models import Sum, Case, When, Value, IntegerField , F
from datetime import datetime
from django.db.models import Count
from matplotlib import pyplot as plt

from .forms import StudentRegistrationForm
from .forms import StudentLoginForm
from .forms import TeacherRegistrationForm
from .forms import TeacherLoginForm
from .forms import ParentRegistrationForm
from .forms import ParentLoginForm
from .forms import AddAttendanceForm

from .models import PARENT
from .models import STUDENT
from .models import STUDENT_INFO
# from .models import TEACHER_INFO
# from .models import SUBJECT
from .models import TEACHER
# from .models import ATTENDANCE_INFO
from .models import ATTENDANCE_DATA

from .resources import PARENTResource
# from .resources import STUDENTResource
# from .resources import STUDENT_INFOResource
from .resources import StudentAndInfoResource
# from .resources import SUBJECTResource
from .resources import TEACHERResource
# from .resources import ATTENDANCEINFOResource
from .resources import ATTENDANCEResource
from django.contrib import messages
from tablib import Dataset

import pandas as pd
import plotly.express as px

from openpyxl import load_workbook
from django.utils import timezone
# from tablib import Dataset

# from .models import ATTENDANCE_INFO
# from .resources import ATTENDANCEResource
# Create or update an instance of YourModel with date_field set to the current date
#-----------------------------------------------------------------
def student_login(request):
        if request.method == 'POST':
            form = StudentLoginForm(request.POST)
            if form.is_valid():
                
                username = form.cleaned_data['username']
                password = form.cleaned_data['password']
                
                if '@' in username:
                    username1 = User.objects.get(email=username.lower()).username
                    user = authenticate(request, username=username1, password=password)

                else:
                    user = authenticate(request, username=username , password=password)
                if user is not None:
                    login(request, user)
                    messages.success(request, "User Login Successful...")
                    return redirect('Student_Dashboard')
                else:
                    messages.info(request,"Invalid username or password !!!")
        else:
            # registerform = StudentRegistrationForm()
            form = StudentLoginForm()
        loginform = form
        return render(request,'student_login.html',{'loginform': loginform })
        
def student_register(request):
        if request.method == 'POST':
            form = StudentRegistrationForm(request.POST)
            if form.is_valid():
                email_id = form.cleaned_data['email_id']
                password = form.cleaned_data['password']
                confirm_password = form.cleaned_data['confirm_password']

                if password == confirm_password:    
                    student = STUDENT.objects.filter(EMAIL_ADDRESS=email_id).first()
                    if student is not None:
                        random_number = random.randint(100, 999)
                        username = student.FIRST_NAME + str(random_number)
                        user = User.objects.create_user(username=username, password=password, email=email_id, first_name=student.FIRST_NAME, last_name=student.LAST_NAME)
                        user.save()
                        messages.success(request, 'User created successfully.')
                        # return redirect('/student_login')  # Redirect to the login page
                    else:
                        messages.error(request, 'Student not found.')
                else:
                    messages.error(request, 'Passwords do not match.')
        else: 
            form = StudentRegistrationForm()
            # loginform = StudentLoginForm()
        registerform = form 
        return render(request,'student_register.html',{'registerform': registerform })

def parent_login(request):
        if request.method == 'POST':
            form = ParentLoginForm(request.POST)
            if form.is_valid():
                
                username = form.cleaned_data['username']
                password = form.cleaned_data['password']
                
                if '@' in username:
                    username1 = User.objects.get(email=username.lower()).username
                    user = authenticate(request, username=username1, password=password)

                else:
                    user = authenticate(request, username=username , password=password)
                if user is not None:
                    login(request, user)
                    messages.success(request, "User Login Successful...")
                    return redirect('Parent_Dashboard')
                else:
                    messages.info(request,"Invalid username or password !!!")
        else:
            # registerform = StudentRegistrationForm()
            form = ParentLoginForm()
        loginform = form
        return render(request,'parent_login.html',{'loginform': loginform })
        
def parent_register(request):
    if request.method == 'POST':
            form = ParentRegistrationForm(request.POST)
            if form.is_valid():
                email_id = form.cleaned_data['email_id']
                password = form.cleaned_data['password']
                confirm_password = form.cleaned_data['confirm_password']

                if password == confirm_password:    
                    parent = PARENT.objects.filter(EMAIL_ADDRESS=email_id).first()
                    if parent is not None:
                        random_number = random.randint(100, 999)
                        username = parent.FIRST_NAME + str(random_number)
                        user = User.objects.create_user(username=username, password=password, email=email_id, first_name=parent.FIRST_NAME, last_name=parent.LAST_NAME)
                        user.save()
                        messages.success(request, 'User created successfully.')
                        # return redirect('/student_login')  # Redirect to the login page
                    else:
                        messages.error(request, 'parent not found.')
                else:
                    messages.error(request, 'Passwords do not match.')
    else: 
            form = ParentRegistrationForm()
            # loginform = StudentLoginForm()
    registerform = form 
    return render(request,'parent_register.html',{'registerform': registerform })

def teacher_login(request):
        if request.method == 'POST':
            form = TeacherLoginForm(request.POST)
            if form.is_valid():
                
                username = form.cleaned_data['username']
                password = form.cleaned_data['password']
                
                if '@' in username:
                    username1 = User.objects.get(email=username.lower()).username
                    user = authenticate(request, username=username1, password=password)

                else:
                    user = authenticate(request, username=username , password=password)
                if user is not None:
                    login(request, user)
                    messages.success(request, "User Login Successful...")
                    return redirect('Teacher_Dashboard')
                else:
                    messages.info(request,"Invalid username or password !!!")
        else:
            # registerform = StudentRegistrationForm()
            form = TeacherLoginForm()
        loginform = form
        return render(request,'teacher_login.html',{'loginform': loginform })

def teacher_register(request):
        if request.method == 'POST':
            form = TeacherRegistrationForm(request.POST)
            if form.is_valid():
                email_id = form.cleaned_data['email_id']
                password = form.cleaned_data['password']
                confirm_password = form.cleaned_data['confirm_password']

                if password == confirm_password:    
                    teacher = TEACHER.objects.filter(EMAIL_ADDRESS=email_id).first()
                    if teacher is not None:
                        random_number = random.randint(100, 999)
                        username = teacher.FIRST_NAME + str(random_number)
                        user = User.objects.create_user(username=username, password=password, email=email_id, first_name=teacher.FIRST_NAME, last_name=teacher.LAST_NAME)
                        user.save()
                        messages.success(request, 'User created successfully.')
                        # return redirect('/student_login')  # Redirect to the login page
                    else:
                        messages.error(request, 'teacher not found.')
                else:
                    messages.error(request, 'Passwords do not match.')
        else: 
            form = TeacherRegistrationForm()
            # loginform = StudentLoginForm()
        registerform = form 
        return render(request,'teacher_register.html',{'registerform': registerform })

#--------------------------------------------------------------------------------

def PARENT_upload(request):
    if request.method == 'POST':
        # Clear existing messages
        storage = messages.get_messages(request)
        storage.used = True

        PARENT_resource = PARENTResource()
        dataset = Dataset()
        new_PARENT = request.FILES['myfile']

        if not new_PARENT.name.endswith('xlsx'):
            messages.info(request,'wrong format')
            return render(request,'attendance_entry.html')

        try:
            imported_data = dataset.load(new_PARENT.read(),format='xlsx')
            for data in imported_data:
                value = PARENT(
                    PARENT_ID=data[0],
                    FIRST_NAME=data[1],
                    LAST_NAME=data[2],
                    PHONE_NO=data[3],
                    EMAIL_ADDRESS=data[4]
                )
                value.save()
            messages.success(request, 'Parent data uploaded successfully.')

        except IntegrityError as ie:
            messages.error(request, f'Data is not unique. Please check your input: {str(ie.args[1])}')
        except ValueError as ve:
            messages.error(request, f'Unsupported data type. Please check your input: {str(ve)}')
        except Exception as e:
            messages.error(request, f'An error occurred: {str(e)}')

    return render(request,'attendance_entry.html')
    
def StudentAndInfo_upload(request):
    # Clear existing messages
    storage = messages.get_messages(request)
    storage.used = True

    if request.method == 'POST':
        STUDENT_resource = StudentAndInfoResource()
        dataset = Dataset()
        new_STUDENT = request.FILES['myfile']

        if not new_STUDENT.name.endswith('xlsx'):
            messages.info(request,'wrong format')
            return render(request,'attendance_entry.html')
        
        try:
            imported_data = dataset.load(new_STUDENT.read(),format='xlsx')
            for data in imported_data:
                parent_id = data[3]  # Assuming data[3] contains the parent_id
                try:
                    parent = PARENT.objects.get(pk=parent_id)
                except PARENT.DoesNotExist:
                    messages.error(request, f"Parent with ID {parent_id} does not exist.")
                    return render(request, 'attendance_entry.html')
                student = STUDENT(
                    STUDENT_ID=data[0],
                    FIRST_NAME=data[1],
                    LAST_NAME=data[2],
                    PARENT_ID=parent,
                    EMAIL_ADDRESS=data[4]
                )
                student.save()
                student_id = data[0]  # Assuming data[3] contains the parent_id
                try:
                    student = STUDENT.objects.get(pk=student_id)
                except STUDENT.DoesNotExist:
                    messages.error(request, f"Student with ID {student_id} does not exist.")
                    return render(request, 'attendance_entry.html')
                student_info = STUDENT_INFO(
                    STUDENT_ID=student,
                    DEPARTMENT=data[5],
                    SECTION=data[6]
                )
                student_info.save()
            messages.success(request, 'Student data uploaded successfully.')
        except IntegrityError as ie:
            messages.error(request, f'Data is not unique. Please check your input: {str(ie.args[1])}')
        except ValueError as ve:
            messages.error(request, f'Unsupported data type. Please check your input: {str(ve)}')
        except Exception as e:
            messages.error(request, f'An error occurred: {str(e)}')

    return render(request,'attendance_entry.html')
    
def TEACHER_upload(request):
    # Clear existing messages
    storage = messages.get_messages(request)
    storage.used = True
    if request.method == 'POST':
        TEACHER_resource = TEACHERResource()
        dataset = Dataset()
        new_TEACHER = request.FILES['myfile']

        if not new_TEACHER.name.endswith('xlsx'):
            messages.info(request,'wrong format')
            return render(request,'attendance_entry.html')
        
        try:
            imported_data = dataset.load(new_TEACHER.read(),format='xlsx')
            for data in imported_data:
                value = TEACHER(
                    TEACHER_ID=data[0],
                    FIRST_NAME=data[1],
                    LAST_NAME=data[2],
                    PHONE_NO=data[3],
                    EMAIL_ADDRESS=data[4]
                )
                value.save()
            messages.success(request, 'Teacher data uploaded successfully.')
        except IntegrityError as ie:
            print(ie.args[1])
            messages.error(request, f'Data is not unique. Please check your input: {str(ie.args[1])}')
        except ValueError as ve:
            messages.error(request, f'Unsupported data type. Please check your input: {str(ve)}')
        except Exception as e:
            messages.error(request, f'An error occurred: {str(e)}')

    return render(request,'attendance_entry.html')
    
def ATTENDANCE_upload(request):
    # Clear existing messages
    storage = messages.get_messages(request)
    storage.used = True
    if request.method == 'POST':
        ATTENDANCEresource = ATTENDANCEResource()
        dataset = Dataset()
        new_ATTENDANCE = request.FILES['myfile']

        if not new_ATTENDANCE.name.endswith('xlsx'):
            messages.info(request,'wrong format')
            return render(request,'attendance_entry.html')
        
        try:
            imported_data = dataset.load(new_ATTENDANCE.read(),format='xlsx')
            for data in imported_data:
                student_id = data[0]
                try:
                    student = STUDENT.objects.get(pk=student_id)
                except STUDENT.DoesNotExist:
                    messages.error(request, f"Student with ID {student_id} does not exist.")
                    return render(request, 'attendance_entry.html')
                value = ATTENDANCE_DATA(
                    STUDENT_ID=student,
                    FIRST_NAME=data[1],
                    LAST_NAME=data[2],
                    DATE=data[3],
                    HOUR1=data[4],
                    HOUR2=data[5],
                    HOUR3=data[6],
                    HOUR4=data[7],
                    HOUR5=data[8],
                    HOUR6=data[9],
                    HOUR7=data[10],
                    HOUR8=data[11],
                )
                value.save()
            messages.success(request, 'Attendance data uploaded successfully.')
        except IntegrityError as ie:
            messages.error(request, f'Data is not unique. Please check your input: {str(ie.args[1])}')
        except ValueError as ve:
            messages.error(request, f'Unsupported data type. Please check your input: {str(ve)}')
        except Exception as e:
            messages.error(request, f'An error occurred: {str(e)}')
    return render(request,'attendance_entry.html')

def attendance_pie(request):
    try:
        attendance_data = ATTENDANCE_DATA.objects.get(STUDENT_ID=20001, DATE="2024-04-03")
    except ATTENDANCE_DATA.DoesNotExist:
        return HttpResponse('Attendance data not found for the given student and date')

    # Count the occurrences of each status (present, absent, od)
    status_counts = {'PRESENT': 0, 'ABSENT': 0, 'ON DUTY': 0}
    for hour in range(1, 9):
        status = getattr(attendance_data, f'HOUR{hour}', None)
        if status == 'PRESENT':
            status_counts['PRESENT'] += 1
        elif status == 'ABSENT':
            status_counts['ABSENT'] += 1
        elif status == 'ON DUTY':
            status_counts['ON DUTY'] += 1

    if not any(status_counts.values()):
        # No valid data for the pie chart
        return HttpResponse('No attendance data available for the given student and date')

    # Plot the data
    plt.figure(figsize=(10, 6))
    plt.pie(status_counts.values(), labels=status_counts.keys(), autopct='%1.1f%%', startangle=90)
    plt.title(f'Attendance Status for Student {20001} on {"2024-04-03"}')
    plt.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle
    plt.savefig('static/plots/attendance.png')  # Save the plot as a PNG file
    plt.close()

    # Pass the path to the saved plot to the template
    context = {'plot_path': 'static/plots/attendance.png'}

    # Render the template with the plot
    return render(request, 'dashboard.html', context)

def overall_attendance(request):
    # Fetch the attendance data
    attendance_data = ATTENDANCE_DATA.objects.all()

    # Convert attendance data to a DataFrame
    attendance_df = pd.DataFrame(list(attendance_data.values()))

    # Calculate total present and absent students for each date
    attendance_df['total_present'] = attendance_df.apply(lambda row: sum(row['HOUR1':'HOUR8'] == 'present'), axis=1)
    attendance_df['total_absent'] = attendance_df.apply(lambda row: sum(row['HOUR1':'HOUR8'] == 'absent'), axis=1)

    # Group by date and calculate total present and absent students for each date
    total_attendance = attendance_df.groupby('DATE').agg({
    'total_present': 'sum',
    'total_absent': 'sum'
    }).reset_index()
    fig = px.line(
        total_attendance,
        x='DATE', 
        y=['total_present', 'total_absent'],

        labels={'date': 'Date', 'value': 'Number of Students', 'variable': 'Attendance Status'},
        title='Overall Attendance Trends')
    chart = fig.to_html()
    context ={'chart':chart}

    return render(request, 'overall_attendance.html', context)
    #return render(request,'overall_attendance.html')
# def download_attendance_excel(request):
#     current_date = timezone.now().date()
#     day_week = current_date.weekday()

#     days={
#          0 : "MONDAY",
#          1 : "TUESDAY",
#          2 : "WEDNESDAY",
#          3 : "THURSDAY",
#          4 : "FRIDAY"
#     }

#     dataset = Dataset()
#     imported_data = dataset.load(open('excels/timetable.xlsx', 'rb').read(),format='xlsx')

#     subjects =[]
#     for data in imported_data:
#         if data[0] == days[day_week]:
#             for i in range(1,9):
#                 subjects.append(data[i])

#     wb = load_workbook('excels/attendance.xlsx')

#     # Iterate over all sheets in the workbook
#     for sheet in wb.sheetnames:
#         ws = wb[sheet]
#         # Delete all rows in the sheet
#         ws.delete_rows(1, ws.max_row)

#     # Define the headers
#     wb.append(["ATTENDANCE_ID","DATE","HOUR","SUBJECT_ID","STATUS"])

#     all_objects = ATTENDANCE_INFO.objects.all()

#     # Iterate over the objects and access their fields
#     for obj in all_objects:
#         for i in range(1,9):
#             wb.append(obj.ATTENDANCE_ID,current_date,i,subjects[i-1],"PRESENT") 

#     wb.save('excels/attendance.xlsx')
#     print("created")    

#-----------------------------------------------------------------------------------------
def home(request):
    return render(request,'index.html')

def login_as(request):
    return render(request,'choicelogin.html')

def teacher_dashboard(request):
    return render(request,'teacher_page.html')

def parent_dashboard(request):
    return render(request,'parent_page.html')

def student_dashboard(request):
    return render(request,'student_page.html')

def dashboard(request):
    return render(request,'dashboard_home.html')

@login_required
def attendance_entry(request):
    return render(request, 'attendance_entry.html')


def add_attendance(request):
    # form = AddAttendanceForm(request.POST or None)
    departments = STUDENT_INFO.objects.values_list('DEPARTMENT', flat=True).distinct()
    sections = STUDENT_INFO.objects.values_list('SECTION', flat=True).distinct()
    if request.method == 'POST':
        # form = AddAttendanceForm(request.POST)
        # if form.is_valid():
        # department = form.cleaned_data['department']
        # section = form.cleaned_data['section']
        # date = form.cleaned_data['date']
        department = request.POST.get('department')
        section = request.POST.get('section')
        date = request.POST.get('date')
        print(department)
        print(section)
        print(date)
        students = STUDENT.objects.filter(student_info__DEPARTMENT=department, student_info__SECTION=section)
        print(students)
        for student in students:
            attendance_data = ATTENDANCE_DATA(
                STUDENT_ID=student,
                FIRST_NAME=student.FIRST_NAME,
                LAST_NAME=student.LAST_NAME,
                DATE=date,
                HOUR1=request.POST.get(f"hour1_{student.STUDENT_ID}"),
                HOUR2=request.POST.get(f"hour2_{student.STUDENT_ID}"),
                HOUR3=request.POST.get(f"hour3_{student.STUDENT_ID}"),
                HOUR4=request.POST.get(f"hour4_{student.STUDENT_ID}"),
                HOUR5=request.POST.get(f"hour5_{student.STUDENT_ID}"),
                HOUR6=request.POST.get(f"hour6_{student.STUDENT_ID}"),
                HOUR7=request.POST.get(f"hour7_{student.STUDENT_ID}"),
                HOUR8=request.POST.get(f"hour8_{student.STUDENT_ID}"),
            )
            attendance_data.save()
        return render(request, 'success.html')
    return render(request, 'add_attendance.html', {'departments': departments, 'sections': sections})



def fetch_students(request):
    department = request.GET.get('department')
    section = request.GET.get('section')
    print("working")
    students = STUDENT.objects.filter(student_info__DEPARTMENT=department, student_info__SECTION=section)
    student_data = []
    for student in students:
        student_data.append({
            'student_id': student.STUDENT_ID,
            'first_name': student.FIRST_NAME,
            'last_name': student.LAST_NAME
        })
    print(student_data)
    
    return JsonResponse(student_data, safe=False)